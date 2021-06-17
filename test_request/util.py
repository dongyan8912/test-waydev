# _*_
# coding: utf-8 _*_


from openpyxl import load_workbook


def read_excel(filename, sheet_name):
    wb = load_workbook(filename)

    sheet = wb[sheet_name]

    request_obj = {}

    for row in sheet.iter_rows(min_row=2, max_row=2, min_col=0, max_col=8):
        request_obj["url"] = row[0].value

        request_obj["app_token"] = row[1].value

        request_obj["app_name"] = row[2].value

        request_obj["activity_id"] = row[3].value

        request_obj["series"] = row[4].value

        request_obj["avatar"] = row[5].value

        request_obj["method"] = row[6].value

    return wb, sheet, row, request_obj


def modify_file(workbook, sheet, name, filename):

    sheet.title = name
    workbook.save(filename)
    workbook.close()

